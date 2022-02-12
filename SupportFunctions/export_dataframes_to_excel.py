import openpyxl
import pandas as pd
import pandas.io.formats.excel
from datetime import date
import openpyxl

from ReportingSourceCode.ReportDefinitions.paths import output_path, template_path, test_output_path, test_template_path


def export_df(df_dict, t_path, o_name) -> str:
    # Clearing formatting of dataframes going into the excel template
    pandas.io.formats.excel.header_style = None
    today = date.today()
    template = openpyxl.load_workbook(test_template_path + t_path)

    with pd.ExcelWriter(test_template_path + t_path, engine='openpyxl') as writer:
        writer.book = template
        writer.sheets = dict((ws.title, ws) for ws in template.worksheets)

        for key in df_dict:
            df_dict[key].to_excel(writer, sheet_name=key, header=False, index=False, startrow=1, startcol=0)

        # Filling the 'Infosheet' date fields
        template['Infosheet']['F11'] = '{}-{}'.format(today.year, today.month)
        template['Infosheet']['F13'] = today

        name_of_report_to_be_saved = o_name + ' {}.xlsx'.format(today)
        template.save(output_path + name_of_report_to_be_saved)

        # Clearing the inserted data in the template
        for key in df_dict:
            for row in template[key]['A2:AF40000']:
                for cell in row:
                    cell.value = None

    return name_of_report_to_be_saved


def export_df_with_header(df_dict, t_path, o_name) -> None:
    # Clearing formatting of dataframes going into the excel template
    pandas.io.formats.excel.header_style = None
    today = date.today()
    template = openpyxl.load_workbook(test_template_path + t_path)

    with pd.ExcelWriter(test_template_path + t_path, engine='openpyxl') as writer:
        writer.book = template
        writer.sheets = dict((ws.title, ws) for ws in template.worksheets)

        for key in df_dict:
            df_dict[key].to_excel(writer, sheet_name=key, header=True, index=False, startrow=0, startcol=0)

        # Filling the 'Infosheet' date fields
        template['Infosheet']['F11'] = '{}-{}'.format(today.year, today.month)
        template['Infosheet']['F13'] = today

        name_of_report_to_be_saved = o_name + ' {}.xlsx'.format(today)
        template.save(output_path + name_of_report_to_be_saved)

        # Clearing the inserted data in the template
        for key in df_dict:
            for row in template[key]['A2:CC40000']:
                for cell in row:
                    cell.value = None

    return name_of_report_to_be_saved


def export_df_with_header_and_index(df_dict, t_path, o_name) -> None:
    # Clearing formatting of dataframes going into the excel template
    pandas.io.formats.excel.header_style = None
    today = date.today()
    template = openpyxl.load_workbook(test_template_path + t_path)

    with pd.ExcelWriter(test_template_path + t_path, engine='openpyxl') as writer:
        writer.book = template
        writer.sheets = dict((ws.title, ws) for ws in template.worksheets)

        for key in df_dict:
            df_dict[key].to_excel(writer, sheet_name=key, header=True, index=True, startrow=0, startcol=0)

        # Filling the 'Infosheet' date fields
        template['Infosheet']['F11'] = '{}-{}'.format(today.year, today.month)
        template['Infosheet']['F13'] = today

        name_of_report_to_be_saved = o_name + ' {}.xlsx'.format(today)
        template.save(output_path + o_name + ' {}.xlsx'.format(today))

        # Clearing the inserted data in the template
        for key in df_dict:
            for row in template[key]['A2:AA40000']:
                for cell in row:
                    cell.value = None

    return name_of_report_to_be_saved


